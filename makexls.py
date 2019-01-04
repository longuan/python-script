# coding:utf-8

from openpyxl import Workbook
# from openpyxl.compat import range
# from openpyxl.utils import get_column_letter

def init_column_1(ws):
    ws.title = "person"

    ws['A1'] = ''
    ws['B1'] = '姓名'
    ws['C1'] = '性别'
    ws['D1'] = '年龄'

def create_workbook(filename):
    wb = Workbook()
    ws1 = wb.active
    init_column_1(ws1)

    a = Person()
    for i in a.make_item():
        # print("writing ", i)
        ws1.append(i)

    wb.save(filename=filename)

if __name__ == "__main__":
    create_workbook("person.xlsx")

# for row in range(1, 40):
#     ws1.append(range(600))

# ws2 = wb.create_sheet(title="Pi")

# ws2['F5'] = 3.14

# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#     for col in range(27, 54):
#         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

# print(ws3['AA10'].value)

# wb.save(filename=dst_filename)
